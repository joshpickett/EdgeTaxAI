from typing import Dict, Any, Union
from datetime import datetime
import io
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from flask import send_file, jsonify

class ExportService:
    def export_pdf(self, data: Dict[str, Any]) -> Union[send_file, tuple]:
        buffer = io.BytesIO()
        pdf = canvas.Canvas(buffer)
        
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawString(50, 800, "Tax Report")
        
        y_position = 750
        pdf.setFont("Helvetica", 12)
        for key, value in data.items():
            pdf.drawString(50, y_position, f"{key}: {value}")
            y_position -= 20
            
        pdf.save()
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'tax_report_{datetime.now().strftime("%Y%m%d")}.pdf'
        )

    def export_excel(self, data: Dict[str, Any]) -> Union[send_file, tuple]:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Tax Report"
        
        for column, header in enumerate(data.keys(), 1):
            worksheet.cell(row=1, column=column, value=header)
        
        for column, value in enumerate(data.values(), 1):
            worksheet.cell(row=2, column=column, value=value)
            
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'tax_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )

    def export_json(self, data: Dict[str, Any]) -> Dict[str, Any]:
        return data
